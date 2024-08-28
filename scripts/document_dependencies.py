#!/usr/bin/env python
import json
import argparse
import openpyxl
from openpyxl.styles import Font, Alignment

ALIGNMENT_DATA = Alignment(wrap_text=False)
ALIGNMENT_HEADER = Alignment(horizontal='center')


def column_number_to_letter(n):
    div = n
    string = ""
    while div > 0:
        module = (div - 1) % 26
        string = chr(65 + module) + string
        div = int((div - module) / 26)
    return string


def vendor_to_excel(row_number, vendor):
    path = vendor["path"]
    revision = vendor["revision"]
    #Vendor
    cell = worksheet.cell(row=row_number, column=1)
    cell.value = path
    cell.alignment = ALIGNMENT_DATA
    #Commit number
    cell = worksheet.cell(row=row_number, column=2)
    cell.value = revision
    cell.alignment = ALIGNMENT_DATA
    #Source code url
    cell = worksheet.cell(row=row_number, column=3)
    cell.value = 'https://'+path+'/archive/'+revision+'.zip'
    cell.hyperlink = 'https://'+path+'/archive/'+revision+'.zip'
    cell.alignment = ALIGNMENT_DATA
    #Community url
    cell = worksheet.cell(row=row_number, column=4)
    cell.value = 'https://'+path
    cell.hyperlink = 'https://'+path
    cell.alignment = ALIGNMENT_DATA
    #License is left to unknown until a license retrieving and parsing method is implemented
    cell = worksheet.cell(row=row_number, column=5)
    cell.value = 'UNKNOWN'
    cell.alignment = ALIGNMENT_DATA



if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generates an Excel file from the vendor list file vendors.json"
    )
    parser.add_argument(
        'origin', metavar='ORIGIN',
        type=argparse.FileType('r'), nargs=1,
        help='The vendor json list')
    parser.add_argument(
        'destination', metavar='DESTINATION',
        type=str, nargs=1,
        help='The output excel file')

    args = parser.parse_args()

    workbook = openpyxl.Workbook()
    workbook.remove_sheet(workbook.worksheets[0])
    worksheet = workbook.create_sheet(title='vendor')

    headers = ["Vendor List", "Revision", "Link to the source code", "Link to the 3PP community", "License"]
    for column_number in xrange(len(headers)):
        cell = worksheet.cell(row=1, column=column_number + 1)
        cell.value = headers[column_number]
        cell.font = Font(bold=True)
        cell.alignment = ALIGNMENT_HEADER

    data = json.load(args.origin[0])
    row_number = 2
    for vendor in data["package"]:
        vendor_to_excel(row_number, vendor)
        row_number += 1

    # The lenght of each column is set to an approximation based on the maximum
    # length of the data in that column. It is an approximation since actual
    # values will be os dependant on runtime
    for column in xrange(len(headers)):
        max_length = 0
        for row in xrange(row_number):
            cell = worksheet.cell(row=row + 1, column=column + 1)
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
        worksheet.column_dimensions[
            column_number_to_letter(column + 1)].width = max_length + 2

    workbook.save(args.destination[0])
